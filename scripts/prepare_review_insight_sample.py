from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def read_sheet_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = [safe_text(v) for v in next(iterator)]
    except StopIteration:
        return []
    rows: list[dict[str, str]] = []
    for row in iterator:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = safe_text(row[idx] if idx < len(row) else "")
        if any(item.values()):
            rows.append(item)
    return rows


def extract_item_id(name: str) -> str:
    match = re.search(r"(\d{10,})", name)
    return match.group(1) if match else ""


def has_media(value: str) -> str:
    text = safe_text(value)
    if not text or text == "[]":
        return "0"
    return "1"


def write_csv(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def pair_files(files: list[Path]) -> list[dict[str, object]]:
    reviews = [p for p in files if "问大家" not in p.name]
    questions = [p for p in files if "问大家" in p.name]
    reviews.sort(key=lambda p: (p.stat().st_mtime, p.name))
    questions.sort(key=lambda p: (p.stat().st_mtime, p.name))

    matches: dict[Path, Path] = {}
    unused_reviews = reviews[:]
    for question_path in questions:
        if not unused_reviews:
            break
        review_path = min(
            unused_reviews,
            key=lambda p: abs(p.stat().st_mtime - question_path.stat().st_mtime),
        )
        matches[review_path] = question_path
        unused_reviews.remove(review_path)

    pairs = []
    for idx, review_path in enumerate(reviews, start=1):
        question_path = matches.get(review_path)
        parent_label = review_path.parent.name if review_path.parent != review_path.parent.parent else ""
        product_label = parent_label if parent_label and parent_label != "00-原始导出" else f"商品样本{idx:03d}"
        pairs.append(
            {
                "sample_id": f"S{idx:03d}",
                "product_label": product_label,
                "item_id": extract_item_id(question_path.name) if question_path else "",
                "review_file": review_path,
                "question_file": question_path,
            }
        )
    return pairs


def main() -> int:
    if len(sys.argv) != 4:
        print(
            "usage: prepare_review_insight_sample.py <raw_dir> <staging_dir> <clean_dir>",
            file=sys.stderr,
        )
        return 2

    raw_dir = Path(sys.argv[1])
    staging_dir = Path(sys.argv[2])
    clean_dir = Path(sys.argv[3])
    files = sorted(raw_dir.rglob("*.xlsx"))
    pairs = pair_files(files)

    inventory_rows = []
    review_rows = []
    question_rows = []
    standard_rows = []

    for pair in pairs:
        sample_id = str(pair["sample_id"])
        product_label = str(pair["product_label"])
        item_id = str(pair["item_id"])
        review_path = pair["review_file"]
        question_path = pair["question_file"]

        review_data = read_sheet_rows(review_path)
        question_data = read_sheet_rows(question_path) if question_path else []

        inventory_rows.append(
            {
                "sample_id": sample_id,
                "product_label": product_label,
                "item_id": item_id,
                "shop_name": "",
                "product_title": "",
                "product_url": "",
                "review_file": review_path.name,
                "review_rows": str(len(review_data)),
                "question_file": question_path.name if question_path else "",
                "question_rows": str(len(question_data)),
                "pairing_basis": "nearest_export_time",
            }
        )

        for idx, row in enumerate(review_data, start=1):
            record = {
                "sample_id": sample_id,
                "product_label": product_label,
                "item_id": item_id,
                "shop_name": "",
                "product_title": "",
                "product_url": "",
                "sku": row.get("SKU", ""),
                "content_type": "review",
                "rating_type": row.get("评价类型", ""),
                "content": row.get("初评", ""),
                "append_content": row.get("追评", ""),
                "has_image": has_media(row.get("晒图/视频", "")),
                "has_video": has_media(row.get("晒图/视频", "")),
                "publish_time": row.get("初评时间", ""),
                "source_file": review_path.name,
                "source_row": str(idx + 1),
                "buyer_name": row.get("旺旺号", ""),
                "helpful_count": row.get("有用", ""),
            }
            review_rows.append(record)
            standard_rows.append(record)

        for idx, row in enumerate(question_data, start=1):
            record = {
                "sample_id": sample_id,
                "product_label": product_label,
                "item_id": item_id,
                "shop_name": "",
                "product_title": "",
                "product_url": "",
                "sku": "",
                "content_type": "question_answer",
                "rating_type": "",
                "content": row.get("问题", ""),
                "append_content": row.get("问答", ""),
                "has_image": "0",
                "has_video": "0",
                "publish_time": row.get("时间", ""),
                "source_file": question_path.name if question_path else "",
                "source_row": str(idx + 1),
                "buyer_name": row.get("昵称", ""),
                "helpful_count": "",
            }
            question_rows.append(record)
            standard_rows.append(record)

    inventory_headers = [
        "sample_id",
        "product_label",
        "item_id",
        "shop_name",
        "product_title",
        "product_url",
        "review_file",
        "review_rows",
        "question_file",
        "question_rows",
        "pairing_basis",
    ]
    standard_headers = [
        "sample_id",
        "product_label",
        "item_id",
        "shop_name",
        "product_title",
        "product_url",
        "sku",
        "content_type",
        "rating_type",
        "content",
        "append_content",
        "has_image",
        "has_video",
        "publish_time",
        "source_file",
        "source_row",
        "buyer_name",
        "helpful_count",
    ]

    write_csv(staging_dir / "商品样本映射表.csv", inventory_rows, inventory_headers)
    write_csv(staging_dir / "评价原始合并.csv", review_rows, standard_headers)
    write_csv(staging_dir / "问大家原始合并.csv", question_rows, standard_headers)
    write_csv(clean_dir / "评价洞察标准数据.csv", standard_rows, standard_headers)

    total_reviews = len(review_rows)
    total_questions = len(question_rows)
    total_rows = len(standard_rows)
    summary = {
        "raw_dir": str(raw_dir),
        "product_samples": len(pairs),
        "review_rows": total_reviews,
        "question_answer_rows": total_questions,
        "standard_rows": total_rows,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "note": "Shop names were not present in source exports. Samples are grouped by export timing and question item_id when available.",
    }
    (staging_dir / "清洗摘要.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    md_lines = [
        "# 评价洞察 Skill 样本01：数据接收与清洗记录",
        "",
        f"生成时间：{summary['generated_at']}",
        "",
        "## 本次接收",
        "",
        f"- 商品样本数：{len(pairs)}",
        f"- 评价记录：{total_reviews}",
        f"- 问大家记录：{total_questions}",
        f"- 标准化总记录：{total_rows}",
        "",
        "## 重要说明",
        "",
        "- 原始导出没有店铺名、商品标题、商品链接，因此先用样本编号和问大家文件里的商品 ID 建立分组。",
        "- 评价表与问大家表按导出时间就近配对，后续如果你能补商品链接或店铺名，可以直接填入 `商品样本映射表.csv`。",
        "- 原始 Excel 没有被修改，全部保留在 `00-原始导出`。",
        "",
        "## 样本映射",
        "",
        "| 样本 | 商品ID | 评价表 | 评价数 | 问大家表 | 问大家数 |",
        "|---|---:|---|---:|---|---:|",
    ]
    for row in inventory_rows:
        md_lines.append(
            f"| {row['sample_id']} | {row['item_id']} | {row['review_file']} | {row['review_rows']} | {row['question_file']} | {row['question_rows']} |"
        )
    md_lines.extend(
        [
            "",
            "## 已生成文件",
            "",
            "- `01-待清洗/商品样本映射表.csv`",
            "- `01-待清洗/评价原始合并.csv`",
            "- `01-待清洗/问大家原始合并.csv`",
            "- `02-已清洗/评价洞察标准数据.csv`",
        ]
    )
    write_md(staging_dir / "数据接收与清洗记录.md", "\n".join(md_lines) + "\n")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
